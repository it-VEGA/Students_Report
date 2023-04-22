from openpyxl import load_workbook
from pathlib import Path

def compile_to_Excel():
    path_to_file = 'teachers-report.txt'
    path = Path(path_to_file)
    wb = load_workbook("data.xlsx")
    ws = wb.active
    ws.title = 'Отсутствующие'
    # ВЫРАВНИВАНИЕ
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30
    #############
    # Запись данных в Excel
    if path.is_file():
        with open("teachers-report.txt", "r", encoding="utf-8") as file:
            src = file.readlines()
            for row in ws.iter_rows(min_row=3, max_col=0, max_row=3):
                for cell in row:
                    for i in range(len(src)):
                        x = src[i].split()
                        cell.value = ws.append(x)
        wb.save("Посещаемость.xlsx")
    else:
        return "Учителя не отправили данные!"
    #######################